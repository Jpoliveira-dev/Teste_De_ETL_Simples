from openpyxl import load_workbook
import pandas as pd

def extrair_produtos():

    arquivo = load_workbook("../data/Controle_Financeiro_Blocos_Concreto.xlsx")
    aba_produtos = arquivo["Produtos"]

    ultima_linha = aba_produtos.max_row

    registros = []

    for linha in range(4, ultima_linha + 1):

        id_produto = aba_produtos.cell(row=linha, column=2).value
        nome_produto = aba_produtos.cell(row=linha, column=3).value
        custo_unitario = aba_produtos.cell(row=linha, column=4).value
        preco_venda = aba_produtos.cell(row=linha, column=5).value
        categoria = aba_produtos.cell(row=linha, column=6).value

        if id_produto == "ID_Produto":
            continue

        if all(v is None for v in [id_produto, nome_produto, custo_unitario, preco_venda, categoria]):
            continue

        registros.append({
            "ID_Produto": id_produto,
            "Nome_Produto": nome_produto,
            "Custo_Unitario": custo_unitario,
            "Preco_Venda_Padrao": preco_venda,
            "Categoria": categoria
        })

    df = pd.DataFrame(registros)

    return df


# teste rápido direto no arquivo
if __name__ == "__main__":
    df = extrair_produtos()
    print(df)
    print("\nTotal de registros:", len(df))
