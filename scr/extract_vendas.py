from openpyxl import load_workbook
import pandas as pd

def extrair_vendas():
    
    arquivo=load_workbook(
        "../data/Controle_Financeiro_Blocos_Concreto.xlsx",
        data_only=True
    )
    
    aba_vendas=arquivo["Vendas"]

    ultima_linha=aba_vendas.max_row

    registros=[]

    for linha in range(4, ultima_linha+1):
        id_venda=aba_vendas.cell(row=linha,column=2).value
        data=aba_vendas.cell(row=linha,column=3).value
        cliente=aba_vendas.cell(row=linha,column=4).value
        id_produto=aba_vendas.cell(row=linha,column=5).value
        quantidade=aba_vendas.cell(row=linha,column=6).value
        preco_venda=aba_vendas.cell(row=linha,column=7).value
        custo_unitario=aba_vendas.cell(row=linha,column=8).value
        receita=aba_vendas.cell(row=linha,column=9).value
        custo_venda=aba_vendas.cell(row=linha,column=10).value
        lucro=aba_vendas.cell(row=linha,column=11).value
        margem=aba_vendas.cell(row=linha,column=12).value
        situacao=aba_vendas.cell(row=linha,column=13).value
        
        if id_venda == "ID_Venda":
            continue

        # if all(v is None for v in [id_venda,lucro,custo_unitario,receita,custo_venda,margem,situacao, data, cliente, id_produto, quantidade, preco_venda]):
        #     continue

        registros.append({
            "ID_Venda":id_venda,
            "Data":data,
            "Cliente":cliente,
            "ID_Produto":id_produto,
            "Quantidade":quantidade,
            "Preço_Venda":preco_venda,
            "Custo_Unitário":custo_unitario,
            "Receita":receita,
            "Custo_Venda":custo_venda,
            "Lucro":lucro,
            "Margem":margem,
            "Situação":situacao,
        })

    df_vendas= pd.DataFrame(registros)

    return df_vendas

if __name__ == "__main__":
     df_vendas = extrair_vendas()
     print(df_vendas)
     print("\nTotal de registros:", len(df_vendas))